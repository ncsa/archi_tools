#!/usr/bin/env python
"""
process an archi .archimate xml into
material needed for re-costing. 

This is a stand-alone tool to generate a
wbs from a .archimate file.

The current processing extracts information from the
folder hierarchy, including descriptions associated
with the folder.

Additional information is generated from information
latent in the folder herarchy. (e.g. wbs numbers)

It would be preferable to process an archimate
standard openexchange file instead of the
archi-version-specific .arcimante file. However,
The .archimate xml file is processed instead of the
arichimate standard openexchange file because the
descriptions associated with folders did not seem
to be exported via the openexcahange format.
Exporting descriptions seem to quite useful in
artifacts generated for downstream tool chains.

"""

import lxml.etree as ET
import argparse
import collections
import db
import shlog

prefix='{http://www.opengroup.org/xsd/archimate/3.0/}'
#tree = ET.parse('LSST.xml')


ALL = []  # hold the list of all folders.

class folderinfo :
    """
    The folderinfo class holds all data extracted from a
    folder as well as data computed per-folder using the
    folder hierarchy (e.g the wbs).  guid name and type
    of elements are collected. 
    """

    HEADER = ["wbs","name","id","documentation","units"]
    def __init__(self, args):
        import collections
        self.args = args
        self.d = collections.defaultdict(lambda : "")
        self.element_list = [] # dictionary holding name, type, id
    def ingest_items(self, items):
        for name, value in items:
            self.d[name]=value
    def ingest_wbslist(self,wbslist):
        self.d["wbs"] = "WBS " + ".".join(wbslist)
    def ingest_documentation(self, documentation):
        self.d["documentation"] = documentation
    def ingest_location(self, location):
        self.d["location"] = location
    def ingest_enclave(self, location):
        self.d["enclave"] = location
    def ingest_element(self, items, element_doc, element_units):
        # provide stucture tuple of items cleaned of XML baggage
        #make a dict to handle the fact that not all things have names
        item_dict = collections.defaultdict(lambda : "")
        item_dict["documentation"] = element_doc
        item_dict["units"] = element_units
        for item in items:
            #handle like this ('{http://www.w3.org/2001/XMLSchema-instance}type', 'archimate:ApplicationComponent')
            if  "type" in item[0]:
                item_dict["type"] = item[1].split(":")[1]
            #handle  this ('name', 'Overall Base AA and ITSEC'),
            elif  "name"  == item[0]:  
                item_dict["name"] = item[1]
            #handle guid like this  ('id', '25c8d1c1-818....')
            elif "id" == item[0]:
                item_dict["id"] = item[1]
        self.element_list.append(item_dict)

    def ingest(self):
        #Do the folder
        insert_list = []
        insert_list.append(self.d['id'])
        insert_list.append(self.d['type'])
        insert_list.append(self.d['name'])
        insert_list.append(self.d['wbs'])
        insert_list.append(self.d['location'])
        insert_list.append(self.d['documentation'])
        insert_list.append(self.d['enclave'])
        insert_list.append(self.d['units'])
        sql = """INSERT INTO FOLDER VALUES (?,?,?,?,?,?,?,?) """        
        db.qp(self.args, sql, [insert_list])
        #make associative table of folder - elements
        sql = "INSERT INTO folder_elements VALUES (?,?)"
        for  element_dict  in self.element_list:
            #import pdb ; pdb.set_trace()
            db.qp(self.args, sql, [[self.d['id'], element_dict['id']]] )
            

    def append_excel(self, worksheet, rowno):
        col = 1
        #make line for folder information
        # source = folder
        # keywords from each
        for h in ["id","wbs","name","documentation","location","enclave"]:
            worksheet.cell(row=rowno, column=col, value=self.d[h])
            worksheet.cell(row=rowno, column=col).font = Font(bold=True)
            worksheet.cell(row=rowno, column=col).alignment = Alignment(wrapText=True)
            col += 1
        rowno += 1
        #  Now lines from nodes in folder
        #  whether to proint the line or continure is continue based on a filter.
        # ginve we are printing a line, cells from the line are computed in differnet ways
        #       repeat some lines from folder  The folder is the "parent" of this stanza
        #       then print some lines from the elment
        #       then for every plateas, print a f(data from element, data from plateau)
        #       idea
        #       functions  fetch via keyword from parent,  f(parent, key)
        #                  fetch from elemnet context      f(self, Key)
        #                  compute from definintion, elmement context) ** is most genreal
        #                                                  f(self, self.context, key) 
        #                      
        #makeline for each element 
        #Hack only report on Nodes and Equipment.
        for element in self.element_list:
            col = 1
            if "Node" in element["type"]  or "Equipment" in element["type"]:
                pass
            else:
                continue
            worksheet.cell(row=rowno, column=col, value=self.d["id"])
            col += 1
            worksheet.cell(row=rowno, column=col, value=self.d["wbs"])
            col += 1
            for item in ["name","documentation","blank","blank","units"] :
                worksheet.cell(row=rowno, column=col, value=element[item])
                worksheet.cell(row=rowno, column=col).alignment = Alignment(wrapText=True)
                col += 1
            #now mark all relevant plateaus
            no_plateaus = True
            sql = "select distinct Pla_name  p from NODE_PLATEAU order by p  Asc"
            for plateau in  db.q(self.args,sql):
                plateau = plateau[0]
                sql = "select count(*) from  NODE_PLATEAU  where Node_id = 'b%s' and Pla_name  =  '%s' "  % (element["id"], plateau)
                count = db.q(self.args, sql).next()[0]
                if count > 0 :
                    worksheet.cell(row=rowno, column=col).alignment = Alignment(wrapText=True)
                    worksheet.cell(row=rowno, column=col, value=plateau.strip())
                    no_plateaus = False
                col +=1
            if no_plateaus :
                red_font = Font(color='00FF0000', italic=True)
                worksheet.cell(row=rowno, column=1).font = red_font
                worksheet.cell(row=rowno, column=2).font = red_font
                worksheet.cell(row=rowno, column=3).font = red_font
                worksheet.cell(row=rowno, column=4).font = red_font
                worksheet.cell(row=rowno, column=5).font = red_font
                worksheet.cell(row=rowno, column=6).font = red_font
            rowno += 1    
        return rowno

    def complete(self):
        ALL.append(self)
        self.ingest()

def get_property (element, property):
    #return the fist key matching property, if no match return empty string.
    #property can be a glob.
    #n.b "property refers to a property attached to an folder, element, or similar in Archimate.
    #a specifc kind fo XML feature related to archimate
    import fnmatch
    
    for e in element.iterchildren("property"):
        for key  in e.values():
            if fnmatch.fnmatch(key, property): return key
    return ""

def wbs(folder, args, wbslist, depth):
    """
    Parse archimate for records related to folders and append
    per-folder folderinfo class objects to the global ALL list.

    Data Model supported:

    In the current archimante (4.2.0) version 
    Each top level archimate folder is is a <folder>
    just under the xml root.  <folders> may contain
    <folders>,  Folders man contain an optional
    <documentation> entity. <folders> contain
    items id and name. e.g.

    <folder name="foldername" id="guid">

"""
    #folders with provisioning estimear have the LDM-129 Property.
    # xml looks like this: <property key="LDM-129"/>
    if get_property(folder,"LDM-129"): 
        info = folderinfo(args)  #container object to load into
        depth=depth+1
    
        #Items are folderitems, like folder names 
        info.ingest_items(folder.items())
        #extract documentions (really there is just one)
        for documentation in  folder.iterchildren("documentation"):
            info.ingest_documentation(documentation.text)
        info.ingest_location(get_property (folder, "Loc*"))
        info.ingest_enclave(get_property (folder, "Enc*"))

        #extract archiment elements 
        for element in  folder.iterchildren("element"):
            #extract documentions (really there is just one)
            doc = ""
            for documentation in  element.iterchildren("documentation"):
                doc = documentation.text
            units = get_property (element, "Unit:*")
            info.ingest_element(element.items(), doc, units)  #list of pairs of items 
            info.ingest_wbslist(wbslist)
        info.complete()
    #recurr over all folders contained in this folder.
    sibno = 1
    for sibling in folder.iterchildren("folder"):
        this_wbslist = wbslist+["%s"%sibno]
        wbs(sibling, args, this_wbslist, depth)  #recur over any children
        sibno += 1        


###########################################################
#
# generic spreadsheet support (nascent stuff to make its
# own library
#
############################################################
        
def make_ws_pretty(args, ws):
    #set columns to reasonable widths.
    colno = 0
    for col in ws.columns:
        colno +=1
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
            adjusted_width = (max_length + 2) * 1.2
            max_width = 60
            if colno > 7 : max_width = 20
            adjusted_width = min (adjusted_width, max_width )
        ws.column_dimensions[column].width = adjusted_width
    return ws


def get_header(args):
    header = ["guid","wbs","element","documentation","location","enclave","unit"]
    sql = "select distinct Pla_name  p from NODE_PLATEAU order by p  Asc"
    for plateau in db.q(args, sql):
        header.append(plateau[0].strip())
    return header
        
###########################################################
#
# Main program
#
############################################################
      

if __name__ == "__main__" :

    import csv
    import sys
    import shlog
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.styles import Alignment
    from openpyxl.styles import PatternFill
    
    #main_parser = argparse.ArgumentParser(add_help=False)
    main_parser = argparse.ArgumentParser(
     description=__doc__,
     formatter_class=argparse.RawDescriptionHelpFormatter,
     fromfile_prefix_chars='@')
    main_parser.add_argument('--loglevel','-l',
                             help='loglevel NONE, NORMAL, VERBOSE, VVERBOSE, DEBUG',
                             default="NONE")
    
    main_parser.add_argument("--show",   "-s", action='store_true', help="pop up excel to show the reult")    
    main_parser.add_argument("--ingest", "-i", action='store_true', help="load teh folders table in the database")    
    main_parser.add_argument("--prefix", "-p", default="LSST_")
    main_parser.add_argument("archimatefile")

    args = main_parser.parse_args()
    shlog.basicConfig(level=shlog.__dict__[args.loglevel])
    args.dbfile = args.prefix + "archi_tool.db"
    shlog.normal ("using database %s", args.dbfile)
    tree = ET.parse(args.archimatefile)
    root = tree.getroot()
    wbs(root, args, [], 0)

    #for row in ALL: row.write_stanza(writer)
    wb = Workbook() #make workbook
    ws = wb.active  #use default sheet
    rowno = 1
    col = 1
    for h in get_header(args):
        ws.cell(row=rowno, column=col).alignment = Alignment(wrapText=True)
        ws.cell(row=rowno, column=col, value=h)
        ws.cell(row=rowno, column=col).font = Font(bold=True)
        col += 1
    rowno = 2
    #build the workbook
    for row in ALL: rowno = row.append_excel(ws, rowno)
    #sets columns to resonable initial width, ets.
    make_ws_pretty(args, ws) 
    wb.save("dog.xlsx")
    import os
    if args.show : os.system('open dog.xlsx -a "Microsoft Excel"')
    
